from openpyxl import load_workbook
from openpyxl.styles import PatternFill

red = PatternFill("solid", "FF0000")


class excel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename, data_only=True)
        self.ws = self.wb[self.sheetname]
