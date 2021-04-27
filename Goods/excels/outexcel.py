from opera.operas import Opera
from goods.good import Good
from goods.node import Node
from copy import deepcopy
from excels.excel import excel, red
from openpyxl import load_workbook


class OutExcel(Opera, excel):
    def __init__(self, filename, sheetname):
        super().__init__(filename, sheetname)
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename, data_only=True)
        self.ws = self.wb[self.sheetname]
        self.node_list = []
        self.good_node = {}
        self.issurewz = 0

    def createDate(self):
        # 必须先获取位置信息
        if self.issurewz == 0:
            self.sureWz()
        # 遍历工作表
        index = self.min_row
        for row in self.ws.iter_rows(min_row=self.min_row, min_col=self.min_col, max_row=self.max_row):
            gd = Good()
            for cell in row:
                cgoodname = self.getColName(self.min_row-2, cell.column)

                if cgoodname is not None and cgoodname != "" and "开通门店" not in cgoodname:
                    result_good = self.fuZhiGood(gd, cell, self.min_row - 2)
                    if not result_good:
                        cell.fill = red
                else:
                    nd = Node()
                    result_node = self.fuZhiNode(nd, cell, self.node_row)
                    if not result_node:
                        self.ws.cell(row=self.node_row, column=cell.column).fill = red
                        del nd
                    else:
                        if nd.isgood != 0:
                            self.node_list.append(nd)

            if gd.check():
                # 保证没个商品都进配中目录
                pz = Node()
                pz.nodecode = "配中"
                self.node_list.append(pz)
                self.good_node[index] = (gd, deepcopy(self.node_list))
            else:
                del gd
            self.node_list.clear()
            index += 1
        self.save(self.filename)
        return self.good_node

    def getColName(self, row, col):
        return self.ws.cell(row=row, column=col).value

    def __getWz(self, value):
        for row in self.ws.iter_rows():
            for cell in row:
                cval = self.valtostring(cell.value)
                if cell.value is not None:
                    if value in cval:
                        return cell.row, cell.column
        return None, None

    def sureWz(self):
        self.node_row, self.node_col = self.__getWz("开通门店")
        self.node_row += 1
        index_num_col = self.__getWz("序号")[1]

        if index_num_col is None:
            self.min_col = 1
        else:
            self.min_col = 2
        # 最小行
        self.min_row = self.node_row + 1
        goodname_col = self.__getWz("商品名称")[1]
        # 最大行
        self.max_row = self.min_row
        while True:
            goodvalue = self.ws.cell(row=self.max_row, column=goodname_col).value
            if goodvalue is None or goodvalue == "":
                if self.max_row - 1 >= self.min_row:
                    self.max_row -= 1
                break
            self.max_row += 1
        self.issurewz = 1

    def save(self, filename):
        self.wb.save(filename)

    def close(self):
        self.wb.close()

    def fuZhiGood(self, gd, cell, row):
        # 必须先获取位置信息
        if self.issurewz == 0:
            self.sureWz()
        # 这里整体可以进行大的优化
        value = cell.value
        # 保证不为None
        colname = Opera().valtostring(self.getColName(row, cell.column))
        if "商品编码" in colname:
            gd.goodscode = value
            return True
        if "是否类别商品" in colname:
            gd.iscategoryCode = value
            if gd.iscategoryCode == "":
                gd.isgood = 0
                return False
            return True
        if "商品类型代码" in colname:
            gd.goodstype = value
            if gd.goodstype == "":
                gd.isgood = 0
                return False
            return True
        if "商品类型名称" in colname:
            gd.goodstypename = value
            return True
        if "分类编码" in colname:
            gd.categorycode = value
            if gd.categorycode == "":
                gd.isgood = 0
                return False
            return True
        if "分类名称" in colname:
            gd.categoryname = value
            if gd.categoryname == "":
                gd.isgood = 0
                return False
            value = self.valtostring(value)
            if gd.categoryname != value:
                gd.isgood = 0
                return False
            return True
        if "商品货号" in colname:
            gd.goodsno = value
            if Opera().countzj(gd.goodsno) > 30:
                return False
            return True
        if "条码" in colname:
            gd.basebarcode = value
            if gd.basebarcode == "":
                return True
            if not gd.basebarcode.isdecimal():
                gd.isgood = 0
                return False
            return True
        if "商品名称（品牌名称+商品名称+规格型号）" in colname:
            gd.goodsname = value
            if gd.goodsname == "":
                gd.isgood = 0
                return False
            return True
        if "商品产地" in colname:
            gd.prodarea = value
            return True
        if "规格型号" in colname:
            gd.goodsspec = value
            return True
        if "计量属性编码" in colname:
            gd.mesureproperty = value
            if gd.mesureproperty == "":
                gd.isgood = 0
                return False
            return True
        if "计量属性对照" in colname:
            gd.mesurepropertyname = value
            if gd.mesurepropertyname == "":
                gd.mesurepropertyname = ""
                gd.isgood = 0
                return False
            return True

        if "商品品牌编码" in colname:
            gd.goodsbrand = value
            return True

        if "商品品牌名称" in colname:
            gd.goodsbrandname = value
            if gd.goodsbrand:
                if gd.goodsbrandname == "":
                    gd.isgood = 0
                    return False
                else:
                    value = self.valtostring(value)
                    if gd.goodsbrandname != value:
                        gd.isgood = 0
                        return False
            return True

        if "进税项" in colname:
            gd.purchtaxrate = value
            if gd.purchtaxrate == "":
                gd.isgood = 0
                return False
            return True
        if "销税项" in colname:
            gd.saletaxrate = value
            if gd.saletaxrate == "":
                gd.isgood = 0
                return False
            return True
        if "基本计量单位" in colname:
            gd.basemeasureunit = value
            return True
        if "供应商编码" in colname:
            gd.suppercode = value
            if gd.suppercode == "":
                gd.isgood = 0
                return False
            return True
        if "供应商合同号编码" in colname:
            gd.cntcode = value
            if gd.cntcode == "":
                gd.isgood = 0
                return False
            return True
        if "保质期" in colname:
            gd.durability = value
            if gd.durability == "":
                gd.isgood = 0
                return False
            return True
        if "小数控制" in colname:
            gd.isdecimalcontrol = value
            if gd.isdecimalcontrol == "":
                gd.isgood = 0
                return False
            return True
        if "进价" == colname:
            gd.purchprice = value
            if gd.purchprice == "":
                gd.isgood = 0
                return False
            return True
        if "售价" == colname:
            gd.saleprice = value
            if gd.saleprice == "":
                gd.isgood = 0
                return False
            return True
        if "海中售价" in colname:
            gd.hzsaleprice = value
            return True
        if "试销期目标金额" in colname:
            gd.specimenmoney = value
            if gd.specimenmoney == "":
                gd.isgood = 0
                return False
            return True
        if "试销天数" in colname:
            gd.specimendays = value
            if gd.specimendays == "":
                gd.isgood = 0
                return False
            return True
        if "整件包装率" in colname:
            gd.wholepackrate = value
            if gd.wholepackrate == "":
                gd.isgood = 0
                return False
            return True
        if "整件包装单位" in colname:
            gd.wholemeasureunit = value
            if gd.wholemeasureunit == "":
                gd.isgood = 0
                return False
            return True
        if "配送加价方式" in colname:
            gd.psjjtype = value
            if gd.psjjtype == "":
                gd.isgood = 0
                return False
            return True
        if "加点率/配送价" in colname:
            gd.jdrate = value
            gd.psj = value
            if gd.jdrate == "" and gd.psj == "":
                gd.isgood = 0
                return False
            return True
        if "门店流转途径" in colname:
            gd.mdlztj = value
            if gd.mdlztj == "":
                gd.isgood = 0
                return False
            return True
        if "整件长" in colname:
            gd.goodslength = value
            return True
        if "整件宽" in colname:
            gd.goodswidth = value
            return True
        if "整件高" in colname:
            gd.goodsheight = value
            return True
        if "自营蚂蚁" in colname or "自营自采" in colname:
            # 通过列名称
            gd.goodspropertycode = colname
            return True
        if "是否退货" in colname:
            gd.isth = value
            return True
        return True

    def fuZhiNode(self, nd, cell, row):
        # 必须先获取位置信息
        if self.issurewz == 0:
            self.sureWz()
        value = cell.value
        colname = self.getColName(row, cell.column)
        if value is not None and value != "":
            nd.nodecode = colname
            if nd.nodecode == "":
                nd.isgood = 0
                return False
            return True
        nd.isgood = 0
        return True

# sqlsv_gbk_gbk = Sqlserve('192.168.8.95', 'sa', '!QAZ2wsx', 'BZ_HQ')
#
# el = OutExcel(r"C:\Users\Administrator\Desktop\提交模版（定）\01 、新品模板(百年版).xlsx", '基本信息（必填）')
# result = el.createDate()
# print(el.min_row, el.min_col, el.max_row, el.node_row, el.node_col)
# for each in result.keys():
#     goodnode = result.get(each)
#     print(each, goodnode[0].goodsname)
#     for node in goodnode[1]:
#         print(node.nodecode, end='\t')
#     print()
# gd = Good()
# print(gd.success_sum, gd.fail_sum)
