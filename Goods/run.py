from excels.outexcel import OutExcel
from excels.indba import InDba
from dba.dba import Sqlserve, sqlsv_gbk, sqlsv_utf, employeename, data_dir, employeecode
from goods.good import Good
import tkinter as tk
from tkinter import Listbox, filedialog, END, ACTIVE
from tkinter import messagebox as mg
import win32com.client as win32
import os
import openpyxl
from versioncheck.versioncheck import Checker
from limit.limit import TimeLimit
from window import window

timelimit = TimeLimit('20210601')

def tLimit(func):
    def inner(*args, **kwargs):
        if timelimit.isOk():
            mg.showwarning(title='提示：', message=f"软件使用时间已到期:王翰卿（Tel：18112992217）")
            window.quit()
            return
        return func(*args, **kwargs)
    return inner


class CreateNewGood:
    fglen = 80

    def __init__(self):
        self.filename = ""
        self.wb = ""
        self.sheetname = ""
        self.outx = ""
        self.indba = ""
        self.result_set = ""
        self.ischeck = 0
        self.iscount = 0

    # 选择worksheet
    @tLimit
    def xzsheet(self, lb):
        # global ws
        self.sheetname = lb.get(ACTIVE)
        print(f"""工作表：{self.sheetname}""")
        print('-' * self.fglen)

    # 创建OutExcel
    @tLimit
    def outExcel(self, filename, sheetname):
        if filename and sheetname:
            self.outx = OutExcel(filename, sheetname)
            return True
        else:
            mg.showwarning(title='提示', message='请打开excel，并选择工作表！')
            print('请打开excel，并选择工作表！')
            print("-" * self.fglen)
            return False

    @tLimit
    def inDba(self):
        self.indba = InDba()

    # 检查excel
    @tLimit
    def checkExcel(self):
        if self.outExcel(self.filename, self.sheetname):
            self.result_set = self.outx.createDate()
            # 判断条码是否重复
            # 判断编码是否重复
            mg.showwarning(title='结果：', message=f'一共有{Good.success_sum + Good.fail_sum}条商品,成功{Good.success_sum}条，'
                                                f'失败{Good.fail_sum}条，请核对Excel！')
            print(f"""工作簿：{self.filename},工作表：{self.sheetname}""")
            print(f'结果：一共有{Good.success_sum + Good.fail_sum}条商品,成功{Good.success_sum}条,失败{Good.fail_sum}条，请核对Excel！')
            print('-' * self.fglen)
            Good.reset()
            self.ischeck = 1

    # 创建基本信息单
    @tLimit
    def goodsBaseBill(self):
        if self.iscount == 0:
            if self.ischeck == 1:
                self.inDba()
                if self.result_set:
                    self.indba.goodsBaseBill(self.result_set, sqlsv_utf)
                    print('-' * self.fglen)
                else:
                    mg.showwarning(title='提示：', message="没有数据！")
                    print("没有数据！")
                    print('-' * self.fglen)
                self.iscount = 1
            else:
                mg.showwarning(title='提示：', message="请先检查Excel！")
                print("请先检查Excel！")
                print('-' * self.fglen)
        else:
            mg.showwarning(title='提示：', message=f"工作簿:{self.filename},\n已经生成基本信息单:{self.indba.BaseBillNumber}")
            print(f"工作簿:{self.filename},\n已经生成基本信息单:{self.indba.BaseBillNumber}")
            print('-' * self.fglen)

    # 进目录
    @tLimit
    def jMl(self, result_set):
        if self.indba == "":
            mg.showwarning(title='注意', message='请先创建基本信息单！')
            print('请先创建基本信息单！')
            print('-' * self.fglen)
        elif self.indba.isExecute(self.indba.BaseBillNumber, sqlsv_gbk):
            self.indba.jMl(result_set, sqlsv_utf)
            print('-' * self.fglen)
        else:
            mg.showwarning(title='提示：', message=f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print(f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print('-' * self.fglen)

    # 配送价格单
    @tLimit
    def pSJGD(self, result_set):
        if self.indba == "":
            mg.showwarning(title='注意', message='请先创建基本信息单！')
            print('请先创建基本信息单！')
            print('-' * self.fglen)
        elif self.indba.isExecute(self.indba.BaseBillNumber, sqlsv_gbk):
            self.indba.pSJGD(result_set, sqlsv_utf)
            mg.showwarning(title='提示：', message=f'请及时执行,配送价格单：{self.indba.psjjBillNumber}')
            print(f'请及时执行,配送价格单：{self.indba.psjjBillNumber}')
            print('-' * self.fglen)
        else:
            mg.showwarning(title='提示：', message=f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print(f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print('-' * self.fglen)

    # 合同商品定义
    @tLimit
    def htGoods(self, result_set):
        if self.indba == "":
            mg.showwarning(title='注意', message='请先创建基本信息单！')
            print('请先创建基本信息单！')
            print('-' * self.fglen)
        elif self.indba.isExecute(self.indba.BaseBillNumber, sqlsv_gbk):
            self.indba.htGoods(result_set, sqlsv_utf)
            mg.showwarning(title='提示：', message=f'请及时执行,合同商品定义单：{self.indba.contractBillNumber}')
            print(f'请及时执行,合同商品定义单：{self.indba.contractBillNumber}')
            print('-' * self.fglen)
        else:
            mg.showwarning(title='提示：', message=f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print(f"请先执行商品基本信息单！{self.indba.BaseBillNumber}")
            print('-' * self.fglen)

    # xls转换成xlsx格式
    @tLimit
    def xls_to_xlsx(self, file_path):
        try:
            print("使用Excel转换！")
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            wb = excel.Workbooks.Open(file_path)
            wb.SaveAs(file_path + 'x', FileFormat=51)
            wb.Close()
            excel.Application.Quit()
        except Exception:
            print("使用WPS转换！")
            wps = win32.gencache.EnsureDispatch('ket.application')
            w = wps.Workbooks.Open(file_path)
            w.SaveAs(file_path + 'x', FileFormat=51)
            w.Close()
            wps.Application.Quit()
        finally:
            print(file_path + ",转化完毕！")
            return file_path + 'x'

    # 打开文件
    @tLimit
    def openfile(self, file, lb):
        self.filename = filedialog.askopenfilename(title='打开文件')
        print(f"""工作簿：{self.filename}""")
        print('-' * self.fglen)
        if self.filename.endswith("xls"):
            file_path_temp = self.filename
            # 将xls转成xlsx
            self.filename = self.xls_to_xlsx(self.filename)
            # 删除原xls文件
            os.remove(file_path_temp)
            file.set(self.filename)
        elif self.filename.endswith('xlsx'):
            file.set(self.filename)
        else:
            mg.showwarning(title='提示：', message="没有选择文件或者文件格式错误，请选择正确的文件格式（xls,xlsx)")
            print("没有选择文件或者文件格式错误，请选择正确的文件格式（xls,xlsx)")
            print('-' * self.fglen)
        try:
            wb = self.wb
            self.wb = openpyxl.load_workbook(self.filename)
            if wb != "":
                # 清理sheet列表类容
                sheetcont = 1
                for wsname in wb.sheetnames:
                    sheetcont += 1
                lb.delete(0, last=sheetcont)
            # 添加sheet列表
            for wsname in self.wb.sheetnames:
                lb.insert(END, wsname)
            self.iscount = 0
        except Exception as e:
            print(e)


def main():
    # 配送价格单，商品是按业态全组织添加
    cnb = CreateNewGood()
    ck = Checker(sqlsv_gbk, '保真连锁超市', '152987')
    success = False
    # 时间到期提示一起
    timec = 0

    # Lable标签显示的类容
    file = tk.StringVar()
    lab = tk.Label(window, text='文件名', bg="red", textvariable=file, width=80, height=2)
    lb = Listbox(window, selectmode='SINGLE')
    b = tk.Button(window, text='打开文件', font=("Arial", 12), width=10, height=1, command=lambda: cnb.openfile(file, lb))
    b2 = tk.Button(window, text='选择', font=("Arial", 12), width=10, height=1, command=lambda: cnb.xzsheet(lb))
    b3 = tk.Button(window, text='检查Excel', font=("Arial", 12), width=10, height=1,
                   command=lambda: cnb.checkExcel())
    b4 = tk.Button(window, text='基本信息单', font=("Arial", 12), width=10, height=1,
                   command=lambda: cnb.goodsBaseBill())
    b5 = tk.Button(window, text="进目录", font=("Arial", 12), width=10, height=1, command=lambda: cnb.jMl(cnb.result_set))
    b6 = tk.Button(window, text="配送价格单", font=("Arial", 12), width=10, height=1,
                   command=lambda: cnb.pSJGD(cnb.result_set))
    b7 = tk.Button(window, text="合同定义", font=("Arial", 12), width=10, height=1,
                   command=lambda: cnb.htGoods(cnb.result_set))

    lab.pack()
    lb.pack()
    b.pack()
    b2.pack()
    b3.pack()
    b4.pack()
    b5.pack()
    b6.pack()
    b7.pack()
    if not success:
        success = ck.cheackVersion()
        if not success:
            mg.showwarning(title='提示：', message=f'该程序是为{ck.COMPANYNAME}公司专有,而不是为贵司订制，如有需要，请联系\n\t王翰卿（Tel：18112992217）')
            return
    if timelimit.getdays() < 2 and timec == 0:
        mg.showwarning(title='提示：', message=f'该程序将在{timelimit.enddate}到期，请及时联系\n\t王翰卿（Tel：18112992217）')
    if not employeename:
        mg.showwarning(title='提示', message=f'配置错误或者数据链接异常!\n您现在链接的服务器是：{data_dir.get("server")}'
                                           f'\n或者工号：{employeecode} ,姓名：{employeename} ,错误'
                                           f'\n请检查数据库配置文件.')
        window.quit()
        return

    window.mainloop()


if __name__ == "__main__":
    main()
